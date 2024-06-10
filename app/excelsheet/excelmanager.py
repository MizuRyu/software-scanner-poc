import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from logging import getLogger

class ExcelHandler:
    def __init__(self, excel_file=None):
        self.logger = getLogger("software-scanner")
        self.excel_file = excel_file or self._default_excel_file_path()

    def _default_excel_file_path(self):
        root = Path(__file__).resolve().parents[2]
        return root / 'data' / 'DeliveryRecords.xlsx'

    def insert_data_to_sheet(self, df, sheet_name):
        self.logger.info(f"Inserting dataframe: {df}")
        try:
            if Path(self.excel_file).exists():
                book = load_workbook(self.excel_file)

                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]

                    start_row = sheet.max_row + 1
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                    # ファイルを保存
                    book.save(self.excel_file)
                    self.logger.info(f"{sheet_name} data inserted successfully.")
                else:
                    self.logger.error(f"Sheet does not exist: {sheet_name}")
                    with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

            else:
                self.logger.error(f"Excel file does not exist: {self.excel_file}")
                with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                self.logger.info(f"{sheet_name} data inserted successfully in new file.")

        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            raise

    def writer_delivery_notes_data(self, delivery_notes_data):
        df = pd.DataFrame(delivery_notes_data)
        self.insert_data_to_sheet(df, "DeliveryNotes")
    
    def writer_product_data(self, product_data_list):
        df = pd.DataFrame(product_data_list)
        self.insert_data_to_sheet(df, "ProductDetails")
    
    def read_data(self):
        try:
            with pd.ExcelFile(self.excel_file) as xls:
                invoice_df = pd.read_excel(xls, sheet_name="DeliveryNotes")
                product_df = pd.read_excel(xls, sheet_name="ProductDetails")
            return invoice_df, product_df
        except Exception as e:
            self.logger.error(f"Error reading data from Excel file: {e}")
            return None, None
